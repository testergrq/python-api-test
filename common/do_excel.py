from openpyxl import load_workbook
from common import contants


class Case:
    def __init__(self):
        self.id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None


class DoExcel:
    # file_name=None

    def __init__(self, file_name):
        try:
            self.file_name = file_name  # 操作的文件
            self.workbook = load_workbook(filename=file_name)  # 实例化一个对象
        except FileNotFoundError as e:# 文件未找到异常处理
            print('{0} not found,please check file path'.format(file_name))
            raise e

    def read_data(self, sheet_name):
        self.sheet_name = sheet_name  #获取表单名
        sheet = self.workbook[sheet_name]  # 定位表单
        max_row = sheet.max_row  # 获取sheet最大行数
        print(max_row)
        cases = []  # 定义一个列表，用来存放即将要放进去的测试用例
        for i in range(2, max_row + 1):
            case = Case()  # 实例化一个case对象，用来存放测试数据
            case.id = sheet.cell(row=i, column=1).value
            case.title = sheet.cell(row=i, column=2).value
            case.url = sheet.cell(row=i, column=3).value
            case.data = sheet.cell(row=i, column=4).value
            case.method = sheet.cell(row=i, column=5).value
            case.expected = sheet.cell(row=i, column=6).value
            if type(case.expected) == int:
                case.expected = str(case.expected)
            cases.append(case) # 将case放到cases 列表里面
        return cases # for 循环结束后返回cases列表

    def write_data(self, sheet_name, row, actual, result):
        sheet = self.workbook[sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.workbook.save(filename=self.file_name)


if __name__ == '__main__':
    do_excel = DoExcel(contants.excel_path)
    cases = do_excel.read_data("login")
    print(cases)
    # request=Request()
    # for case in cases:
    #     print('开始执行第{0}条用例'.format(case.id))
    #     resp=request.request(case.method,case.url,case.data)
    #     # print(resp.json())
    #     if resp.text == case.expected:
    #         do_excel.write_data(case.id+1,resp.text,'PASS')
    #         print('第{0}条用例执行结果:pass'.format(case.id))
    #     else:
    #         do_excel.write_data(case.id+1,resp.text,'FAIL')
    #         print('第{0}条用例执行结果:fail'.format(case.id))
